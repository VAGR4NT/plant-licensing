from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, Http404
from django.urls import reverse
from django import forms
from django.contrib.contenttypes.models import ContentType
from django.forms import modelform_factory
from django.db import connection
from .query_builder import (
    MODEL_MAP,
    query_builder,
    format_table,
    _validate_and_get_field,
)
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.http import FileResponse, HttpResponse
import json
from django.apps import apps
from django.db.models import ForeignKey, OneToOneField, ManyToManyField
from django.views.generic import CreateView, UpdateView
from django.db.models import F
from .models import EmailTemplate
from django.contrib import messages
from io import BytesIO
from pathlib import Path
from django.conf import settings
from django.utils import timezone

from .models import RenewalTemplate
from .forms import TemplateUploadForm

import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Protection, PatternFill

from pypdf import PdfReader, PdfWriter
from pypdf.generic import NameObject, BooleanObject, DictionaryObject

from .models import Businesses, Locations, Licenses, Suppliers, ComplianceAgreements
from .forms import DealerForm, NurseryForm, LocationForm, ComplianceForm, SupplierForm
from django.db import transaction
from .models import Businesses, Locations, Licenses, Suppliers

from django.core.files import File

import base64
import uuid
from email.utils import formatdate


def home_view(request):
    return render(request, "main/index.html", context)


def specific_view(request):
    table_fields_data = {}

    significant_tables = {
        "Business Details": {"model": Businesses, "prefix": ""},
        "Location Details": {"model": Locations, "prefix": "locations__"},
        "License Details": {"model": Licenses, "prefix": "locations__licenses__"},
        "Supplier Details": {"model": Suppliers, "prefix": "suppliers__"},
    }

    for table, info in significant_tables.items():
        fields_for_group = []
        model_class = info["model"]
        path_prefix = info["prefix"]

        for field in model_class._meta.fields:
            if not isinstance(field, (ForeignKey, OneToOneField)):
                clean_label = field.name.replace("_", " ").title()
                fields_for_group.append(
                    {"name": f"{path_prefix}{field.name}", "label": clean_label}
                )
        table_fields_data[table] = fields_for_group

    context = {
        "table_fields": table_fields_data,
        "operators": ["exact", "iexact", "icontains", "gt", "gte", "lt", "lte"],
        "headers": [],
        "rows": [],
        "business_model_id": ContentType.objects.get_for_model(Businesses).id,
        "error_message": None,
        "result_count": None,
        "request_get": request.GET,
        "checked_columns": request.GET.getlist("columns"),
    }

    try:
        filters = []
        if request.GET.get("q"):
            filters.append(
                {
                    "field": "business_name",
                    "operator": "icontains",
                    "value": request.GET.get("q"),
                }
            )

        for i in range(1, 6):
            field = request.GET.get(f"adv_field_{i}")
            op = request.GET.get(f"adv_op_{i}")
            val = request.GET.get(f"adv_val_{i}")
            if field and op and val:
                filters.append({"field": field, "operator": op, "value": val})

        select_columns = request.GET.getlist("columns")
        if not select_columns:
            select_columns = ["business_name", "mo_city"]

        queryset = query_builder(Businesses, filters=filters)
        context["result_count"] = queryset.count()

        display_data = format_table(queryset, select_columns)
        context.update(display_data)

    except ValueError as e:
        context["error_message"] = f"Error building query: {e}"

    return render(request, "main/view/specific_view.html", context)


def independent_view(request, ct):
    model_class = ContentType.objects.get(id=ct).model_class()

    context = {
        "headers": [],
        "rows": [],
        "model_id": ct,
        "model_name": model_class._meta.verbose_name.title(),
        "add_url": model_class.get_add_url()
        if hasattr(model_class, "get_add_url")
        else None,
        "error_message": None,
    }

    try:
        filters = []

        queryset = query_builder(model_class, filters=filters)

        print(queryset)
        print(model_class.identifying_fields())

        display_data = format_table(queryset, model_class.identifying_fields())
        context.update(display_data)
        print(context)

    except ValueError as e:
        context["error_message"] = f"Error building query: {e}"

    return render(request, "main/view/independent_view.html", context)


def direct_access_view(request):
    return render(request, "main/direct-access/index.html")


def view_db_view(request):
    context = {
        "independent_models": {},
    }

    django_junk = [
        "DjangoAdminLog",
        "DjangoContentType",
        "DjangoMigrations",
        "DjangoSession",
        "AuthUser",
        "AuthGroup",
        "Businesses",
        "EmailTemplate",
        "RenewalTemplate",
        "SellingSeasons",
    ]

    try:
        # Get a list of all our tables
        all_models = apps.get_app_config("main").get_models()

        # Loop through each model and check if it is dependent
        for model in all_models:
            if model._meta.object_name in django_junk:
                continue

            is_independent = True
            for field in model._meta.fields:
                if isinstance(field, (ForeignKey, OneToOneField)):
                    is_independent = False
                    break

            if is_independent:
                context["independent_models"][model._meta.verbose_name.title()] = (
                    ContentType.objects.get_for_model(model).id
                )

    except LookupError:
        context["error"] = "Problem"

    return render(request, "main/view/index.html", context)


def generate_forms_view(request):
    return render(request, "main/generate-forms/index.html")


class add_business(CreateView):
    model = Businesses
    template_name = "main/add_business/index.html"

    def get_form_class(self):
        form_type = self.request.GET.get("type", "Dealer")

        if form_type == "Nursery":
            return NurseryForm
        return DealerForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context["current_type"] = self.request.GET.get("type", "Dealer")
        return context

    def get_success_url(self):
        return reverse(
            "update",
            kwargs={
                "ct": ContentType.objects.get_for_model(self.object._meta.model).id,
                "pk": self.object.pk,
            },
        )


class add_compliance_agreement(CreateView):
    model = ComplianceAgreements
    template_name = "main/add_business/add_location.html"

    def get_form_class(self):
        return ComplianceForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["parent_id"] = self.kwargs.get("parent_id")
        return kwargs

    def get_success_url(self):
        return reverse(
            "update",
            kwargs={
                "ct": ContentType.objects.get_for_model(self.object._meta.model).id,
                "pk": self.object.pk,
            },
        )


class add_supplier(CreateView):
    model = Suppliers
    template_name = "main/add_business/add_independent.html"

    def get_form_class(self):
        return SupplierForm

    def get_success_url(self):
        return reverse(
            "update",
            kwargs={
                "ct": ContentType.objects.get_for_model(self.object._meta.model).id,
                "pk": self.object.pk,
            },
        )

    def form_valid(self, form):
        response = super().form_valid(form)

        if "popup" in self.request.GET:
            return render(
                self.request,
                "main/add_business/close_popup.html",
                {"object": self.object},
            )

        return response


class add_location(CreateView):
    model = Locations
    template_name = "main/add_business/add_location.html"

    def get_form_class(self):
        return LocationForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["parent_id"] = self.kwargs.get("parent_id")
        return kwargs

    def get_success_url(self):
        return reverse(
            "update",
            kwargs={
                "ct": ContentType.objects.get_for_model(self.object._meta.model).id,
                "pk": self.object.pk,
            },
        )


def update_view(request, ct, pk):
    """
    A dynamic view to display and update any model instance ("master_object")
    and list its related "child" objects.
    """
    model_class = ContentType.objects.get(id=ct).model_class()
    master_object = get_object_or_404(model_class, pk=pk)
    local_field_names = [
        field.name for field in model_class._meta.fields if not field.is_relation
    ]
    DynamicModelForm = modelform_factory(model_class, fields=local_field_names)

    if request.method == "POST":
        if "delete" in request.POST:
            master_object.delete()
            return redirect(request.META.get("HTTP_REFERER", "/"))

        form = DynamicModelForm(request.POST, instance=master_object)

        if form.is_valid():
            form.save()
            return redirect(request.path)

    else:
        form = DynamicModelForm(instance=master_object)

    related_sections = {}

    all_related_fields = model_class._meta.get_fields(include_hidden=False)

    for field in all_related_fields:
        if field.is_relation and (field.one_to_many):
            accessor_name = field.get_accessor_name()
            if not accessor_name or not hasattr(master_object, accessor_name):
                continue

            queryset = getattr(master_object, accessor_name).all()
            related_model = field.related_model
            related_sections[related_model._meta.verbose_name.title()] = {
                "items": queryset,
                "model": ContentType.objects.get_for_model(related_model).id,
                "add_url": related_model.get_add_url()
                if hasattr(related_model, "get_add_url")
                else None,
            }

    context = {
        "master_object": master_object,
        "form": form,
        "related_sections": related_sections,
        "master_model_name": model_class._meta.model_name,
    }

    if hasattr(master_object, "get_parent"):
        context["parent_object"] = master_object.get_parent()
        context["parent_ct"] = ContentType.objects.get_for_model(
            master_object.get_parent()
        ).id
    else:
        context["parent_object"] = None
        context["parent_ct"] = None

    if hasattr(master_object, "get_add_url"):
        context["add_url"] = master_object.get_add_url()
    else:
        context["add_url"] = None

    return render(request, "main/update/index.html", context)


def user_info_view(request):
    return render(request, "main/user-info/index.html")


def account_view(request):
    return render(request, "main/account/index.html")


# ============================================================
# SHARED PDF FILL FUNCTION
# ============================================================


def _fill_pdf(template_path: str, field_map: dict) -> bytes:
    reader = PdfReader(template_path)

    # Guard: template must have AcroForm
    root = reader.trailer.get("/Root", {})
    acroform = root.get("/AcroForm")
    if not acroform:
        raise ValueError(
            f"{template_path} has no AcroForm (not a fillable PDF). "
            "Ensure you are using a fillable template."
        )

    writer = PdfWriter()

    # Copy pages
    for p in reader.pages:
        writer.add_page(p)

    # Copy AcroForm
    writer._root_object.update({NameObject("/AcroForm"): acroform})

    # Apply field values
    writer.update_page_form_field_values(writer.pages[0], field_map)

    # Ensure appearance streams are rendered
    writer._root_object["/AcroForm"].update(
        {NameObject("/NeedAppearances"): BooleanObject(True)}
    )

    out = BytesIO()
    writer.write(out)
    return out.getvalue()


# ============================================================
# NURSERY FIELD HELPERS
# ============================================================


def _compute_amount_due(acreage: float) -> tuple[str, float]:
    amount = 40.00 + 1.50 * acreage
    return f"${amount:,.2f}", amount


def _extract_location_lines(biz: Businesses, max_lines: int = 4):
    """Return (list_of_lines, first_location_obj)."""
    first_loc = Locations.objects.filter(business=biz).order_by("location_id").first()
    lines = []

    # Prefer notes
    if first_loc and (first_loc.field_location_notes or "").strip():
        for line in (first_loc.field_location_notes or "").splitlines():
            stripped = line.strip()
            if stripped:
                lines.append(stripped)
            if len(lines) == max_lines:
                break
    else:
        # Fallback: city/county/zip
        qs = (
            Locations.objects.filter(business=biz)
            .order_by("location_id")
            .values("city", "county", "zip_code")
        )[:max_lines]

        for rec in qs:
            parts = [
                p
                for p in [rec.get("city"), rec.get("county"), rec.get("zip_code")]
                if p
            ]
            if parts:
                lines.append(", ".join(parts))

    while len(lines) < max_lines:
        lines.append("")

    return lines, first_loc


def _nursery_fieldmap(biz: Businesses):
    acreage = float(biz.acreage or 0)
    amount_due_str, _ = _compute_amount_due(acreage)
    location_lines, first_loc = _extract_location_lines(biz)

    checkbox_val = "Yes"
    today_str = timezone.localdate().strftime("%m/%d/%Y")

    return {
        "License Number": str(biz.pk),
        "current_license_year1": "",
        "business_name": biz.business_name or "",
        "mailing_address": biz.mo_address or "",
        "main_office_city1": biz.mo_city or "",
        "main_office_state": biz.mo_state or "",
        "main_office_zip1": biz.mo_zip or "",
        "main_office_county": (
            first_loc.county if first_loc and first_loc.county else ""
        ),
        "main_contact_name": biz.main_contact_name or "",
        "main_office_city2": biz.mo_city or "",
        "main_office_zip2": biz.mo_zip or "",
        "phone_number": biz.main_contact_phone or "",
        "fax": "",
        "main_contact_email": biz.main_contact_email or "",
        "field_location1": location_lines[0],
        "field_location2": location_lines[1],
        "field_location3": location_lines[2],
        "field_location4": location_lines[3],
        "acreage": f"{acreage:g}",
        "amount_due": amount_due_str,
        "Amt": amount_due_str,
        "Check Box17": checkbox_val,
        "Date": today_str,
        "Lic Year": "",
        "Date Recd": "",
        "Check": "",
    }

def download_nursery_pdf(request, business_id: int):
    # Get the business row by business_id (this is your “key”)
    try:
        biz = Businesses.objects.get(pk=business_id)
    except Businesses.DoesNotExist:
        raise Http404("Business not found")

    # Choose a representative location (adjust selection rule if needed)
    location = Locations.objects.filter(business=biz).order_by("location_id").first()

    # Fees: $40 + $1.50 per acre
    acreage = float(biz.acreage or 0)
    amount_due_val = 40.00 + 1.50 * acreage
    amount_due_str = f"${amount_due_val:,.2f}"

    # Field locations: split notes across up to 4 lines (or synthesize)
    location_lines = []
    if location and (location.field_location_notes or "").strip():
        for line in (location.field_location_notes or "").splitlines():
            line = line.strip()
            if line:
                location_lines.append(line)
            if len(location_lines) == 4:
                break
    else:
        qs = (
            Locations.objects.filter(business=biz)
            .order_by("location_id")
            .values("city", "county", "zip_code")
        )[:4]
        for rec in qs:
            parts = [
                p
                for p in [rec.get("city"), rec.get("county"), rec.get("zip_code")]
                if p
            ]
            if parts:
                location_lines.append(", ".join(parts))
    while len(location_lines) < 4:
        location_lines.append("")

    # Checkbox (email vs USPS) — set as you wish or leave "Off"
    prefers_email = True
    checkbox_val = "Yes" if prefers_email else "Off"

    # Today’s date for the certification line
    today_str = timezone.localdate().strftime("%m/%d/%Y")

    # Map DB → THIS form’s fields
    field_map = {
        # Header: put business_id into “License Number”
        "License Number": str(business_id),
        "current_license_year1": "",
        # Mailing Information
        "business_name": biz.business_name or "",
        "mailing_address": biz.mo_address or "",
        "main_office_city1": biz.mo_city or "",
        "main_office_state": biz.mo_state or "",
        "main_office_zip1": biz.mo_zip or "",
        "main_office_county": (location.county if location and location.county else ""),
        # Contact
        "main_contact_name": biz.main_contact_name or "",
        "main_office_city2": biz.mo_city or "",
        "main_office_zip2": biz.mo_zip or "",
        "phone_number": biz.main_contact_phone or "",
        "fax": "",
        "main_contact_email": biz.main_contact_email or "",
        # Field locations (4 lines on this form)
        "field_location1": location_lines[0],
        "field_location2": location_lines[1],
        "field_location3": location_lines[2],
        "field_location4": location_lines[3],
        # Fees
        "acreage": f"{acreage:g}",
        "amount_due": amount_due_str,
        "Amt": amount_due_str,  # mirror into bottom box if desired
        # Checkbox + date
        "Check Box17": checkbox_val,
        "Date": today_str,
        # Bottom admin box
        "Lic Year": "",
        "Date Recd": "",
        "Check": "",
    }

    filled = _fill_pdf(str(TEMPLATE_PATH), field_map)
    resp = HttpResponse(filled, content_type="application/pdf")
    resp["Content-Disposition"] = (
        f'attachment; filename="nursery_renewal_{business_id}.pdf"'
    )
    return resp


def download_nursery_pdf(request, business_id: int):
    # Get the business row by business_id (this is your “key”)
    try:
        biz = Businesses.objects.get(pk=business_id)
    except Businesses.DoesNotExist:
        raise Http404("Business not found")

    # Choose a representative location (adjust selection rule if needed)
    location = Locations.objects.filter(business=biz).order_by("location_id").first()

    # Fees: $40 + $1.50 per acre
    acreage = float(biz.acreage or 0)
    amount_due_val = 40.00 + 1.50 * acreage
    amount_due_str = f"${amount_due_val:,.2f}"

    # Field locations: split notes across up to 4 lines (or synthesize)
    location_lines = []
    if location and (location.field_location_notes or "").strip():
        for line in (location.field_location_notes or "").splitlines():
            line = line.strip()
            if line:
                location_lines.append(line)
            if len(location_lines) == 4:
                break
    else:
        qs = (
            Locations.objects.filter(business=biz)
            .order_by("location_id")
            .values("city", "county", "zip_code")
        )[:4]
        for rec in qs:
            parts = [
                p
                for p in [rec.get("city"), rec.get("county"), rec.get("zip_code")]
                if p
            ]
            if parts:
                location_lines.append(", ".join(parts))
    while len(location_lines) < 4:
        location_lines.append("")

    # Checkbox (email vs USPS) — set as you wish or leave "Off"
    prefers_email = True
    checkbox_val = "Yes" if prefers_email else "Off"

    # Today’s date for the certification line
    today_str = timezone.localdate().strftime("%m/%d/%Y")

    # Map DB → THIS form’s fields
    field_map = {
        # Header: put business_id into “License Number”
        "License Number": str(business_id),
        "current_license_year1": "",
        # Mailing Information
        "business_name": biz.business_name or "",
        "mailing_address": biz.mo_address or "",
        "main_office_city1": biz.mo_city or "",
        "main_office_state": biz.mo_state or "",
        "main_office_zip1": biz.mo_zip or "",
        "main_office_county": (location.county if location and location.county else ""),
        # Contact
        "main_contact_name": biz.main_contact_name or "",
        "main_office_city2": biz.mo_city or "",
        "main_office_zip2": biz.mo_zip or "",
        "phone_number": biz.main_contact_phone or "",
        "fax": "",
        "main_contact_email": biz.main_contact_email or "",
        # Field locations (4 lines on this form)
        "field_location1": location_lines[0],
        "field_location2": location_lines[1],
        "field_location3": location_lines[2],
        "field_location4": location_lines[3],
        # Fees
        "acreage": f"{acreage:g}",
        "amount_due": amount_due_str,
        "Amt": amount_due_str,  # mirror into bottom box if desired
        # Checkbox + date
        "Check Box17": checkbox_val,
        "Date": today_str,
        # Bottom admin box
        "Lic Year": "",
        "Date Recd": "",
        "Check": "",
    }

    filled = _fill_pdf(str(TEMPLATE_PATH), field_map)
    resp = HttpResponse(filled, content_type="application/pdf")
    resp["Content-Disposition"] = (
        f'attachment; filename="nursery_renewal_{business_id}.pdf"'
    )
    return resp


# ============================================================
# DEALER FIELDMAP
# ============================================================


def _dealer_fieldmap(biz: Businesses):
    return {
        "business_name": biz.business_name,
        "mailing_address": biz.mo_address,
        "main_office_city": biz.mo_city,
        "main_office_state": biz.mo_state,
        "main_office_zip": biz.mo_zip,
        "main_contact_name": biz.main_contact_name or "",
        "main_contact_phone": biz.main_contact_phone or "",
        "fax": "",
        "email": biz.main_contact_email or "",
        "renewal_year1": "",
        "renewal_year2": "",
        "current_license_year": "",
        "email_pref": True,
        "usps_pref": False,
    }


# ============================================================
# EML BUILDER
# ============================================================


def _build_eml(
    to_addr: str, subject: str, body: str, pdf_bytes: bytes, filename: str
) -> str:
    boundary = uuid.uuid4().hex
    pdf_b64 = base64.b64encode(pdf_bytes).decode("ascii")

    return f"""X-Unsent: 1
From:
To: {to_addr}
Subject: {subject}
Date: {formatdate(localtime=True)}
MIME-Version: 1.0
Content-Type: multipart/mixed; boundary="{boundary}"

--{boundary}
Content-Type: text/plain; charset="utf-8"

{body}

--{boundary}
Content-Type: application/pdf
Content-Disposition: attachment; filename="{filename}"
Content-Transfer-Encoding: base64

{pdf_b64}

--{boundary}--
"""


LICENSE_TYPES = {
    "nursery": {
        "template_defaults": {
            "subject": "Plant License Renewal for {business_name}",
            "body": 'Attached is the renewal form: "{filename}".',
        },
        "fieldmap": _nursery_fieldmap,
    },
    "dealer": {
        "template_defaults": {
            "subject": "Plant License Renewal for {business_name}",
            "body": 'Attached is the renewal form: "{filename}".',
        },
        "fieldmap": _dealer_fieldmap,
    },
}


def _get_license_config(kind: str):
    try:
        return LICENSE_TYPES[kind]
    except KeyError:
        raise Http404(f"Unknown license type: {kind}")


def _load_email_template(kind: str):
    cfg = _get_license_config(kind)
    return EmailTemplate.objects.get_or_create(
        name=kind,
        defaults=cfg["template_defaults"],
    )


# ============================================================
# SHARED PDF/EML GENERATORS
# ============================================================


def generate_pdf(kind: str, biz: Businesses):
    cfg = _get_license_config(kind)

    field_map = cfg["fieldmap"](biz)

    template_path = _get_template_path(kind)
    pdf_bytes = _fill_pdf(template_path, field_map)

    filename = f"{kind} renewal - {biz.business_name}.pdf"
    return pdf_bytes, filename


def preview_pdf(request, kind: str, business_id: int):
    biz = get_object_or_404(Businesses, pk=business_id)
    pdf_bytes, filename = generate_pdf(kind, biz)

    resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    resp["Content-Length"] = len(pdf_bytes)
    return resp


def download_pdf(request, kind: str, business_id: int):
    biz = get_object_or_404(Businesses, pk=business_id)
    pdf_bytes, filename = generate_pdf(kind, biz)

    resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def download_eml(request, kind: str, business_id: int):
    """
    Generate an EML email containing the PDF as an attachment.
    Supports full {field_name} placeholders anywhere in the template,
    including all business fields and first location fields.
    """
    biz = get_object_or_404(Businesses, pk=business_id)
    tmpl, _ = _load_email_template(kind)

    # Generate PDF
    pdf_bytes, filename = generate_pdf(kind, biz)

    # Start with PDF field map
    field_map = LICENSE_TYPES[kind]["fieldmap"](biz)
    field_map["filename"] = filename

    # Merge all fields from the Businesses model
    for field in Businesses._meta.get_fields():
        if hasattr(biz, field.name):
            field_map[field.name] = getattr(biz, field.name)

    # Merge all fields from the first location
    first_loc = Locations.objects.filter(business=biz).order_by("location_id").first()
    if first_loc:
        for field in Locations._meta.get_fields():
            if hasattr(first_loc, field.name):
                field_map[field.name] = getattr(first_loc, field.name)

    # Interpolate placeholders in subject and body
    subject = tmpl.subject.format(**field_map)
    body = tmpl.body.format(**field_map)

    # Build raw EML
    eml_body = _build_eml(
        to_addr=biz.main_contact_email or "",
        subject=subject,
        body=body,
        pdf_bytes=pdf_bytes,
        filename=filename,
    )

    resp = HttpResponse(eml_body, content_type="message/rfc822")
    resp["Content-Disposition"] = f'attachment; filename="{filename}.eml"'
    return resp


# ============================================================
# SHARED GENERATE PAGES (nursery & dealer)
# ============================================================


def generate_page(request, kind: str, template_name: str):
    tmpl, _ = _load_email_template(kind)

    # --- Handle email template save ---
    if request.method == "POST":
        tmpl.subject = request.POST.get("subject", "")
        tmpl.body = request.POST.get("body", "")
        tmpl.save()
        messages.success(request, "Email template updated.")

    # Toggle filter based on query parameter
    show_wants = request.GET.get("show_wants", "1") == "1"  # default True

    businesses = Businesses.objects.filter(wants_email_renewal=show_wants).order_by(
        "business_name"
    )

    if reverse_order:
        # reversed: False first, then True
        businesses = Businesses.objects.order_by(
            F("wants_email_renewal").asc(nulls_last=True),
            "business_name",
        )
    else:
        # normal: True first, then False
        businesses = Businesses.objects.order_by(
            F("wants_email_renewal").desc(nulls_last=True),
            "business_name",
        )

    # --- Render page ---
    return render(
        request,
        template_name,
        {
            "businesses": businesses,
            "template": tmpl,
            "show_wants": show_wants,
        },
    )


def nursery_generate(request):
    return generate_page(request, "nursery", "main/nursery_generate/index.html")


def dealer_generate(request):
    return generate_page(request, "dealer", "main/dealer_generate/index.html")


DEFAULT_PDFS = {
    "nursery": "plant_license/pdfs/nursery_renewal_fillable.pdf",
    "dealer": "plant_license/pdfs/dealer_renewal_fillable.pdf",
}


def pdf_update(request):
    if request.method == "POST":
        kind = request.POST.get("kind")
        form = TemplateUploadForm(request.POST, request.FILES)

        if form.is_valid():
            uploaded_file = request.FILES["pdf"]  # assuming field is named pdf
            with transaction.atomic():
                # deactivate old active
                RenewalTemplate.objects.filter(kind=kind, is_active=True).update(
                    is_active=False
                )

                # save new template
                new_template = form.save(commit=False)
                new_template.kind = kind
                new_template.is_active = True
                new_template.original_filename = uploaded_file.name
                new_template.save()

            messages.success(request, f"{kind.title()} PDF updated successfully.")
            return redirect("pdf_update")

    return render(
        request,
        "main/pdf_update/index.html",
        {
            "nursery_form": TemplateUploadForm(),
            "dealer_form": TemplateUploadForm(),
            "nursery_templates": RenewalTemplate.objects.filter(kind="nursery"),
            "dealer_templates": RenewalTemplate.objects.filter(kind="dealer"),
        },
    )


@transaction.atomic
def remove_template(request, template_id: int):
    template = get_object_or_404(RenewalTemplate, id=template_id)

    if template.is_active:
        messages.error(
            request, "Cannot remove the active template. Activate another first."
        )
        return redirect("pdf_update")

    template.delete()
    messages.success(request, f"{template.kind.title()} template removed successfully.")
    return redirect("pdf_update")


def _get_template_path(kind: str) -> str:
    template = RenewalTemplate.objects.filter(kind=kind, is_active=True).first()
    if template:
        if not template.original_filename:
            template.original_filename = template.pdf.name.split("/")[-1]
            template.save()
        return template.pdf.path

    if kind not in DEFAULT_PDFS:
        raise Http404(f"No default PDF configured for {kind}")

    default_rel_path = DEFAULT_PDFS[kind]

    default_abs_path = finders.find(default_rel_path)
    if not default_abs_path:
        raise Http404(f"Default PDF missing for {kind}: {default_rel_path}")

    default_abs_path = Path(default_abs_path)

    pdf_bytes = default_abs_path.read_bytes()
    filename = default_abs_path.name

    stored_path = default_storage.save(
        f"pdfs/{filename}",
        ContentFile(pdf_bytes),
    )

    template = RenewalTemplate.objects.create(
        kind=kind,
        pdf=stored_path,
        original_filename=filename,
        is_active=True,
    )

    return template.pdf.path


def download_template_version(request, template_id: int):
    template = get_object_or_404(RenewalTemplate, pk=template_id)

    return FileResponse(
        template.pdf.open("rb"),
        as_attachment=True,
        filename=template.original_filename,
    )


@transaction.atomic
def activate_template_version(request, template_id: int):
    template = get_object_or_404(RenewalTemplate, id=template_id)

    # deactivate all others of same kind
    RenewalTemplate.objects.filter(kind=template.kind).update(is_active=False)

    # activate selected
    template.is_active = True
    template.save()

    messages.success(request, f"{template.kind.title()} template is now active.")

    return redirect("pdf_update")


def export_table_as_csv(request):
    # Create the HttpResponse object with the appropriate CSV header.
    # This tells the browser to treat the response as a file to be downloaded.
    response = HttpResponse(
        content_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="data.csv"'},
    )

    # Get all data from your model
    # This assumes your model is SampleData. Replace with your actual model.
    queryset = Businesses.objects.all()

    # Check if there is any data to write
    if not queryset.exists():
        response.write("No data found.")
        return response

    # Get the model's field names
    # We use _meta.fields to get all field objects
    # and then get the name for each field.
    field_names = [field.name for field in Businesses._meta.fields]

    # Create a CSV writer object using the response as the file
    writer = csv.writer(response)

    # Write the header row
    writer.writerow(field_names)

    # Iterate over the queryset and write each row to the CSV
    for obj in queryset:
        # Create a list of values for the current object
        row = [getattr(obj, field) for field in field_names]
        writer.writerow(row)

    return response


def export_table_as_xlsx(request):
    # 1. Create a new Workbook and get the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Business Data"

    PROTECTION_PASSWORD = "LockHeader"

    # 2. Configure the HTTP Response for XLSX
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="data_protected.xlsx"'},
    )

    # 3. Get all data and field names
    queryset = Businesses.objects.all()
    if not queryset.exists():
        pass  # Handle empty queryset as before

    field_names = [field.name for field in Businesses._meta.fields]
    num_columns = len(field_names)

    # 4. Write the header row (This row will remain locked by default)
    worksheet.append(field_names)

    # 5. Write data rows AND UNLOCK the data cells

    # We start iterating from row 2 (index 1) for data
    row_num = 2
    for obj in queryset:
        row_data = [getattr(obj, field) for field in field_names]
        worksheet.append(row_data)

        # Unlock all cells in the current data row (row_num)
        for col_index in range(1, num_columns + 1):
            col_letter = get_column_letter(col_index)
            cell = worksheet[f"{col_letter}{row_num}"]

            # Key Step 1: Set the protection to 'unlocked'
            cell.protection = Protection(locked=False)

        row_num += 1

    worksheet.protection.sheet = True
    worksheet.protection.password = PROTECTION_PASSWORD

    # You can configure what the user is allowed to do while the sheet is protected
    worksheet.protection.sort = True
    worksheet.protection.autofilter = True
    worksheet.protection.insertRows = True
    worksheet.protection.insertColumns = True

    # 7. Save the workbook to the HttpResponse file-like object
    workbook.save(response)

    return response
